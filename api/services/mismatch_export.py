from __future__ import annotations

import csv
import io
from typing import Any

from fastapi.responses import StreamingResponse

from api.services.run_report import RunReportSnapshot
from etl_framework.utils.serialization import csv_safe

MISMATCH_FIELDS = [
    "test_name",
    "key_values",
    "column_name",
    "source_value",
    "target_value",
    "mismatch_type",
]


def collect_mismatch_rows(repo: Any, snapshot: RunReportSnapshot) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for result in snapshot.results:
        if result.id is None:
            continue
        for mismatch in repo.list_mismatches(result_id=result.id, limit=100_000):
            rows.append({
                "test_name": result.query_name,
                "key_values": csv_safe(mismatch.key_values),
                "column_name": mismatch.column_name or "",
                "source_value": mismatch.source_value or "",
                "target_value": mismatch.target_value or "",
                "mismatch_type": mismatch.mismatch_type or "",
            })
    return rows


def mismatch_csv_response(run_id: str, rows: list[dict[str, str]]) -> StreamingResponse:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=MISMATCH_FIELDS)
    writer.writeheader()
    writer.writerows(rows)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="mismatches_{run_id}.csv"'},
    )


def mismatch_xlsx_response(
    run_id: str,
    snapshot: RunReportSnapshot,
    rows: list[dict[str, str]],
) -> StreamingResponse:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    fills = {
        "value_diff": PatternFill("solid", fgColor="FEF3C7"),
        "value_mismatch": PatternFill("solid", fgColor="FEF3C7"),
        "missing_in_target": PatternFill("solid", fgColor="FEE2E2"),
        "missing_in_source": PatternFill("solid", fgColor="EDE9FE"),
    }
    header_fill = PatternFill("solid", fgColor="1E293B")
    header_font = Font(bold=True, color="F1F5F9")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mismatches"
    ws.append(MISMATCH_FIELDS)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:F1"

    for row_data in rows:
        ws.append([row_data[field] for field in MISMATCH_FIELDS])
        fill = fills.get(row_data["mismatch_type"])
        if fill:
            for cell in ws[ws.max_row]:
                cell.fill = fill

    for col, width in [("A", 30), ("B", 28), ("C", 22), ("D", 30), ("E", 30), ("F", 22)]:
        ws.column_dimensions[col].width = width

    _append_summary_sheet(wb, run_id, snapshot, header_fill, header_font)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="mismatches_{run_id}.xlsx"'},
    )


def _append_summary_sheet(wb: Any, run_id: str, snapshot: RunReportSnapshot, header_fill: Any, header_font: Any) -> None:
    ws = wb.create_sheet("Summary")
    ws.title = "Summary"
    ws.append(["Run ID", run_id])
    ws.append(["Status", snapshot.status])
    ws.append(["Raw Status", snapshot.raw_status])
    ws.append(["Started", str(snapshot.started_at or "")])
    ws.append(["Completed", str(snapshot.completed_at or "")])
    ws.append(["Source Env", snapshot.source_env or ""])
    ws.append(["Target Env", snapshot.target_env or ""])
    ws.append([])
    ws.append([
        "Test Name",
        "Status",
        "Raw Status",
        "Source Rows",
        "Target Rows",
        "Value Mismatches",
        "Missing in Target",
        "Missing in Source",
    ])
    for cell in ws[ws.max_row]:
        cell.fill = header_fill
        cell.font = header_font
    for result in snapshot.results:
        ws.append([
            result.query_name,
            result.effective_status,
            result.status,
            result.source_row_count,
            result.target_row_count,
            result.value_mismatch_count,
            result.missing_in_target_count,
            result.missing_in_source_count,
        ])
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 16


