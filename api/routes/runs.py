from __future__ import annotations

import csv
import asyncio
import io
import json
import time
import uuid
from collections.abc import Callable
from datetime import datetime, timezone
from pathlib import Path

from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException, Request
from fastapi.responses import PlainTextResponse, FileResponse, HTMLResponse, StreamingResponse
from pydantic import ValidationError
from sqlalchemy.orm import Session

from api.dependencies import get_session
from api.routes.configs import _preserve_masked_secrets
from api.schemas import (
    DrilldownOut,
    DrilldownRequest,
    DrilldownRow,
    GeneratedArtifactOut,
    MismatchAcceptOut,
    MismatchAcceptRequest,
    MismatchOut,
    RunCompareOut,
    RunDetailOut,
    RunProgressOut,
    RunStatusOut,
    RunStepOut,
    RunStepReleaseRequest,
    RunTrigger,
    TestCompareOut,
    TestResultOut,
    TestResultOverrideRequest,
    TestSuiteTrigger,
)
from api.services.run_executor import RunExecutor
from api.services.pytest_runner import PytestRunExecutor
from etl_framework.repository.repository import ConfigRepository, RunRepository, RunStepRepository
from api.services.artifact_service import ArtifactService
from api.services.artifact_views import render_logs_html, render_metrics_html
from api.services.audit_service import AuditService
from api.services.log_parser import detect_log_level, parse_log_events, filter_log_events
from etl_framework.config.models import resolve_connection as _resolve_connection
from etl_framework.repository.models import TERMINAL_STATUSES as _TERMINAL

router = APIRouter(tags=["runs"])


_TREND_CACHE_TTL_SECONDS = 30
_TREND_CACHE: dict[tuple, tuple[float, dict]] = {}


def _test_result_out(result) -> TestResultOut:
    return TestResultOut(
        id=result.id,
        query_name=result.query_name,
        status=result.status,
        effective_status=result.effective_status,
        duration_seconds=result.duration_seconds,
        source_row_count=result.source_row_count,
        target_row_count=result.target_row_count,
        value_mismatch_count=result.value_mismatch_count,
        missing_in_target_count=result.missing_in_target_count,
        missing_in_source_count=result.missing_in_source_count,
        error_message=result.error_message,
        executed_at=result.executed_at,
        override_reason=result.override_reason,
        overridden_by=result.override_by,
        override_at=result.override_at,
        sample_rows=result.sample_rows,
        segment_summary=result.segment_summary,
    )


def _wants_html(request: Request, fmt: str | None) -> bool:
    if fmt:
        return fmt.lower() == "html"
    user_agent = request.headers.get("user-agent", "").lower()
    if "testclient" in user_agent:
        return False
    return "text/html" in request.headers.get("accept", "").lower()




def _metrics_from_run(run) -> dict:
    tests = []
    total_duration = 0.0
    passed = failed = slow = 0
    for result in run.results:
        duration = float(result.duration_seconds or 0)
        total_duration += duration
        status = getattr(result, "effective_status", None) or result.status or "UNKNOWN"
        if status == "PASSED":
            passed += 1
        elif status == "SLOW":
            slow += 1
        elif status in {"FAILED", "ERROR"}:
            failed += 1
        tests.append({
            "name": result.query_name,
            "status": status,
            "duration_seconds": duration,
            "source_row_count": result.source_row_count or 0,
            "target_row_count": result.target_row_count or 0,
            "total_issues": result.total_issues,
        })
    return {
        "run_id": run.run_id,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "total_tests": len(tests),
        "passed": passed,
        "failed": failed,
        "slow": slow,
        "total_duration_seconds": round(total_duration, 6),
        "tests": tests,
        "source": "database",
    }


def _validate_connection_name(cfg, name: str | None, field: str) -> None:
    if name is None or cfg is None:
        return
    available = list((cfg.config_json or {}).get("connections", {}).keys())
    if name not in available:
        raise HTTPException(
            status_code=422,
            detail={
                "message": f"{field} '{name}' not found in config connections",
                "available": available,
            },
        )


def _resolve_connection_or_422(source: dict, name: str | None, env_name: str, field: str):
    try:
        return _resolve_connection(source, name, env_name=env_name)
    except ValidationError as exc:
        raise HTTPException(
            status_code=422,
            detail={"message": f"{field} '{name}' failed validation", "errors": exc.errors()},
        ) from exc


def _snapshot_from_trigger(body: RunTrigger, db: Session) -> dict:
    cfg_data = dict(body.config_data or {})
    cfg = ConfigRepository(db).get(body.config_id) if body.config_id is not None else None
    if body.config_id is not None and cfg is None:
        raise HTTPException(status_code=404, detail="Config not found")
    if cfg is not None:
        # config_data may be an echo of the masked GET /api/configs response
        # (e.g. the Launch page's Saved Config dropdown) — restore the real
        # stored secret wherever the client sent back the display mask.
        cfg_data = _preserve_masked_secrets(cfg_data, cfg.config_json)
        cfg_data = {**(cfg.config_json or {}), **cfg_data}

    snapshot = dict(cfg_data)
    if cfg is not None:
        snapshot.update({
            "config_id": cfg.id,
            "config_name": cfg.name,
            "env_name": cfg.env_name,
        })

    if "source_credentials" not in snapshot:
        _validate_connection_name(cfg, body.source_connection, "source_connection")
        if body.source_connection:
            src = _resolve_connection_or_422(
                cfg.config_json if cfg else cfg_data,
                body.source_connection,
                body.source_env,
                "source_connection",
            )
            snapshot["source_credentials"] = {**src.model_dump(), "name": body.source_env}
        else:
            snapshot["source_credentials"] = {
                "name": body.source_env,
                **{k: v for k, v in cfg_data.items() if k != "connections"},
            }
    if "target_credentials" not in snapshot:
        _validate_connection_name(cfg, body.target_connection, "target_connection")
        if body.target_connection:
            tgt = _resolve_connection_or_422(
                cfg.config_json if cfg else cfg_data,
                body.target_connection,
                body.target_env,
                "target_connection",
            )
            snapshot["target_credentials"] = {**tgt.model_dump(), "name": body.target_env}
        else:
            snapshot["target_credentials"] = {
                "name": body.target_env,
                **{k: v for k, v in cfg_data.items() if k != "connections"},
            }
    if "bo_credentials" not in snapshot:
        snapshot["bo_credentials"] = {"name": "bo", **cfg_data}
    if "automic_credentials" not in snapshot:
        snapshot["automic_credentials"] = {"name": "automic", **cfg_data}
    return snapshot


def _execute_run(
    run_id: str,
    job_sequence: list[str],
    source_env: str,
    target_env: str,
    run_settings,
    config_snapshot: dict | None,
    session_factory: Callable[[], Session] | None = None,
) -> None:
    from etl_framework.repository.database import SessionLocal
    from etl_framework.utils.context import set_run_id

    set_run_id(run_id)

    db = (session_factory or SessionLocal)()
    try:
        RunExecutor(
            db=db,
            run_id=run_id,
            source_env=source_env,
            target_env=target_env,
            job_sequence=job_sequence,
            run_settings=run_settings,
            config_snapshot=config_snapshot,
        ).execute()
    finally:
        db.close()


def _run_pytest(
    run_id: str,
    pytest_args: list[str],
    session_factory=None,
) -> None:
    from etl_framework.repository.database import SessionLocal
    from etl_framework.utils.context import set_run_id

    set_run_id(run_id)
    db = (session_factory or SessionLocal)()
    try:
        PytestRunExecutor(db=db, run_id=run_id, pytest_args=pytest_args).execute()
    finally:
        db.close()


@router.get("", response_model=list[RunStatusOut])
def list_runs(
    limit: int = 50,
    offset: int = 0,
    status: str | None = None,
    run_type: str | None = None,
    db: Session = Depends(get_session),
):
    repo = RunRepository(db)
    runs = repo.list_runs(limit=limit, offset=offset, status=status, run_type=run_type)
    return [
        RunStatusOut(
            run_id=r.run_id,
            status=r.status,
            started_at=r.started_at,
            completed_at=r.completed_at,
            total_tests=r.total_tests,
            passed=r.passed,
            failed=r.failed,
            slow=r.slow,
            error=r.error,
            run_type=r.run_type,
            pair_id=r.pair_id,
        )
        for r in runs
    ]


@router.post("/test-suite", response_model=RunStatusOut, status_code=202)
def trigger_test_suite(
    body: TestSuiteTrigger,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_session),
):
    run_id = str(uuid.uuid4())
    RunRepository(db).create_run(
        run_id=run_id,
        source_env=None,
        target_env=None,
        run_type="test_suite",
    )
    background_tasks.add_task(_run_pytest, run_id, body.pytest_args)
    return RunStatusOut(run_id=run_id, status="PENDING", run_type="test_suite")


@router.post("", response_model=RunStatusOut, status_code=202)
def trigger_run(
    body: RunTrigger,
    background_tasks: BackgroundTasks,
    request: Request,
    db: Session = Depends(get_session),
):
    run_id = str(uuid.uuid4())
    repo = RunRepository(db)
    ordered_jobs = body.job_sequence or body.job_names
    run_settings = body.run_settings.model_dump()
    config_snapshot = _snapshot_from_trigger(body, db)
    if ordered_jobs:
        # Serialize SequenceStep objects to plain dicts for JSON storage
        config_snapshot["job_sequence"] = [
            s.model_dump() if hasattr(s, "model_dump") else s
            for s in ordered_jobs
        ]
    config_snapshot["run_settings"] = run_settings
    repo.create_run(
        run_id=run_id,
        source_env=body.source_env,
        target_env=body.target_env,
        config_snapshot=config_snapshot or None,
    )
    AuditService(db).log(
        request,
        "run.created",
        "run",
        run_id,
        {
            "source_env": body.source_env,
            "target_env": body.target_env,
            "job_sequence": config_snapshot.get("job_sequence"),
            "config_id": body.config_id,
        },
    )
    background_tasks.add_task(
        _execute_run,
        run_id,
        ordered_jobs,
        body.source_env,
        body.target_env,
        body.run_settings,
        config_snapshot,
    )
    return RunStatusOut(run_id=run_id, status="PENDING")


@router.post("/{run_id}/cancel", status_code=202)
def cancel_run(run_id: str, db: Session = Depends(get_session)):
    repo = RunRepository(db)
    if repo.get_run(run_id) is None:
        raise HTTPException(status_code=404, detail="Run not found")
    accepted = repo.request_cancel(run_id)
    return {"run_id": run_id, "cancel_requested": accepted}


@router.get("/{run_id}/status", response_model=RunStatusOut)
def get_run_status(run_id: str, db: Session = Depends(get_session)):
    repo = RunRepository(db)
    run = repo.get_run(run_id)
    if run is None:
        raise HTTPException(status_code=404, detail="Run not found")
    return RunStatusOut(
        run_id=run.run_id,
        status=run.status,
        started_at=run.started_at,
        completed_at=run.completed_at,
        total_tests=run.total_tests,
        passed=run.passed,
        failed=run.failed,
        slow=run.slow,
        error=run.error,
        run_type=run.run_type,
        pair_id=run.pair_id,
    )


@router.get("/{run_id}/artifacts", response_model=list[GeneratedArtifactOut])
def list_run_artifacts(run_id: str, db: Session = Depends(get_session)):
    if RunRepository(db).get_run(run_id) is None:
        raise HTTPException(status_code=404, detail="Run not found")
    artifacts = []

    artifacts.append(
        GeneratedArtifactOut(
            name=f"report_{run_id}.html",
            artifact_type="report",
            path=f"/api/runs/{run_id}/report",
            created_at=datetime.now(timezone.utc)
        )
    )

    metrics_path = Path("logs") / f"metrics_{run_id}.json"
    if metrics_path.exists():
        artifacts.append(
            GeneratedArtifactOut(
                name=metrics_path.name,
                artifact_type="metrics",
                path=f"/api/runs/{run_id}/metrics",
                created_at=datetime.fromtimestamp(
                    metrics_path.stat().st_mtime, tz=timezone.utc
                ),
            )
        )
    log_path = Path("logs") / "etl_framework.log"
    if log_path.exists():
        artifacts.append(
            GeneratedArtifactOut(
                name=log_path.name,
                artifact_type="log",
                path=f"/api/runs/{run_id}/logs",
                created_at=datetime.fromtimestamp(
                    log_path.stat().st_mtime, tz=timezone.utc
                ),
            )
        )
    return artifacts


@router.get("/{run_id}/metrics")
def get_run_metrics(
    run_id: str,
    request: Request,
    format: str | None = None,
    db: Session = Depends(get_session),
):
    run = RunRepository(db).get_run(run_id)
    if run is None:
        raise HTTPException(status_code=404, detail="Run not found")
    metrics_path = Path("logs") / f"metrics_{run_id}.json"
    if metrics_path.exists():
        metrics = json.loads(metrics_path.read_text(encoding="utf-8", errors="replace"))
    elif run.results:
        metrics = _metrics_from_run(run)
    else:
        raise HTTPException(status_code=404, detail="Metrics not found")
    if _wants_html(request, format):
        return HTMLResponse(render_metrics_html(metrics))
    return metrics


@router.get("/{run_id}/logs")
def get_run_logs(
    run_id: str,
    request: Request,
    q: str = "",
    level: str = "",
    limit: int = 500,
    scope: str = "run",
    format: str | None = None,
    db: Session = Depends(get_session),
):
    if RunRepository(db).get_run(run_id) is None:
        raise HTTPException(status_code=404, detail="Run not found")
    log_path = Path("logs") / "etl_framework.log"
    if not log_path.exists():
        raise HTTPException(status_code=404, detail="Log not found")
    text = log_path.read_text(encoding="utf-8", errors="replace")
    scope_l = scope.lower().strip() or "run"
    if scope_l not in {"run", "all"}:
        raise HTTPException(status_code=400, detail="scope must be 'run' or 'all'")
    total_events = len(parse_log_events(text))
    lines = filter_log_events(
        text,
        run_id=run_id if scope_l == "run" else "",
        query=q,
        level=level,
        limit=limit,
    )
    fmt = (format or "").lower()
    if fmt == "json":
        return {
            "run_id": run_id,
            "query": q,
            "level": level,
            "scope": scope_l,
            "total_lines": len(text.splitlines()),
            "total_events": total_events,
            "matched_lines": len(lines),
            "lines": lines,
        }
    if _wants_html(request, format):
        return HTMLResponse(render_logs_html(
            run_id=run_id,
            lines=lines,
            query=q,
            level=level.upper().strip(),
            total_lines=len(text.splitlines()),
            total_events=total_events,
            scope=scope_l,
        ))
    if q or level or scope_l == "run":
        return PlainTextResponse("\n".join(row["text"] for row in lines))
    return PlainTextResponse(text)


@router.get("/{run_id}/report", response_class=FileResponse)
def get_run_report(run_id: str, db: Session = Depends(get_session)):
    service = ArtifactService(repository=RunRepository(db))
    report_path = service.generate_html_report(run_id)
    return FileResponse(report_path, media_type="text/html")


@router.get("/{run_id}/progress", response_model=RunProgressOut)
def get_run_progress(run_id: str, db: Session = Depends(get_session)):
    repo = RunRepository(db)
    run = repo.get_run(run_id)
    if run is None:
        raise HTTPException(status_code=404, detail="Run not found")
    total = run.total_tests or 0
    completed = repo.count_completed_results(run_id)
    percent = int(completed / total * 100) if total > 0 else 0
    return RunProgressOut(
        run_id=run.run_id,
        status=run.status,
        total_tests=total,
        completed_tests=completed,
        current_job=repo.get_current_job(run_id),
        percent_complete=min(percent, 100),
    )


@router.get("/{run_id}/results/{result_id}/mismatches", response_model=list[MismatchOut])
def list_result_mismatches(
    run_id: str,
    result_id: int,
    limit: int = 100,
    offset: int = 0,
    db: Session = Depends(get_session),
):
    repo = RunRepository(db)
    if repo.get_run(run_id) is None:
        raise HTTPException(status_code=404, detail="Run not found")
    rows = repo.list_mismatches(result_id=result_id, limit=limit, offset=offset)
    return [
        MismatchOut(
            id=m.id,
            column_name=m.column_name,
            key_values=m.key_values,
            source_value=m.source_value,
            target_value=m.target_value,
            mismatch_type=m.mismatch_type,
            accepted=m.accepted,
            accepted_note=m.accepted_note,
            accepted_at=m.accepted_at,
            accepted_by=m.accepted_by,
        )
        for m in rows
    ]


@router.get("/{run_id}/mismatches/download")
def download_mismatches(
    run_id: str,
    format: str = "csv",
    db: Session = Depends(get_session),
):
    """Download all mismatch details for a run as CSV, XLSX, or HTML report."""
    repo = RunRepository(db)
    run = repo.get_run(run_id)
    if run is None:
        raise HTTPException(status_code=404, detail="Run not found")

    if format == "html":
        service = ArtifactService(repository=repo)
        report_path = service.generate_html_report(run_id)
        return FileResponse(
            report_path,
            media_type="text/html",
            headers={"Content-Disposition": f'attachment; filename="report_{run_id}.html"'},
        )

    _FIELDS = ["test_name", "key_values", "column_name", "source_value", "target_value", "mismatch_type"]
    rows = []
    for result in run.results:
        for m in repo.list_mismatches(result_id=result.id, limit=100_000):
            rows.append({
                "test_name": result.query_name,
                "key_values": json.dumps(m.key_values) if isinstance(m.key_values, dict) else str(m.key_values or ""),
                "column_name": m.column_name or "",
                "source_value": m.source_value or "",
                "target_value": m.target_value or "",
                "mismatch_type": m.mismatch_type or "",
            })

    if format == "xlsx":
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment

        FILLS = {
            "value_diff":        PatternFill("solid", fgColor="FEF3C7"),  # amber
            "value_mismatch":    PatternFill("solid", fgColor="FEF3C7"),  # amber
            "missing_in_target": PatternFill("solid", fgColor="FEE2E2"),  # rose
            "missing_in_source": PatternFill("solid", fgColor="EDE9FE"),  # violet
        }
        HEADER_FILL = PatternFill("solid", fgColor="1E293B")
        HEADER_FONT = Font(bold=True, color="F1F5F9")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Mismatches"
        ws.append(_FIELDS)
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = "A1:F1"

        for row_data in rows:
            ws.append([row_data[field] for field in _FIELDS])
            fill = FILLS.get(row_data["mismatch_type"])
            if fill:
                for cell in ws[ws.max_row]:
                    cell.fill = fill

        for col, width in [("A", 30), ("B", 28), ("C", 22), ("D", 30), ("E", 30), ("F", 22)]:
            ws.column_dimensions[col].width = width

        # ── Summary sheet ────────────────────────────────────────────────
        ws_sum = wb.create_sheet("Summary")
        ws_sum.title = "Summary"
        ws_sum.append(["Run ID", run_id])
        ws_sum.append(["Status", run.status])
        ws_sum.append(["Started", str(run.started_at or "")])
        ws_sum.append(["Completed", str(run.completed_at or "")])
        ws_sum.append(["Source Env", run.source_env or ""])
        ws_sum.append(["Target Env", run.target_env or ""])
        ws_sum.append([])
        ws_sum.append(["Test Name", "Status", "Source Rows", "Target Rows",
                        "Value Mismatches", "Missing in Target", "Missing in Source"])
        for cell in ws_sum[ws_sum.max_row]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        for result in run.results:
            ws_sum.append([
                result.query_name,
                result.status,
                result.source_row_count,
                result.target_row_count,
                result.value_mismatch_count,
                result.missing_in_target_count,
                result.missing_in_source_count,
            ])
        ws_sum.column_dimensions["A"].width = 36
        ws_sum.column_dimensions["B"].width = 16

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="mismatches_{run_id}.xlsx"'},
        )

    buf_str = io.StringIO()
    writer = csv.DictWriter(buf_str, fieldnames=_FIELDS)
    writer.writeheader()
    writer.writerows(rows)
    buf_str.seek(0)
    return StreamingResponse(
        iter([buf_str.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="mismatches_{run_id}.csv"'},
    )


@router.patch(
    "/{run_id}/results/{result_id}/mismatches/{mismatch_id}/accept",
    response_model=MismatchAcceptOut,
)
def accept_mismatch(
    run_id: str,
    result_id: int,
    mismatch_id: int,
    body: MismatchAcceptRequest,
    request: Request,
    db: Session = Depends(get_session),
):
    from etl_framework.repository.models import MismatchDetail, TestResult

    repo = RunRepository(db)
    if repo.get_run(run_id) is None:
        raise HTTPException(status_code=404, detail="Run not found")
    tr = db.get(TestResult, result_id)
    if tr is None or tr.run_id != run_id:
        raise HTTPException(status_code=404, detail="Result not found")
    md = db.get(MismatchDetail, mismatch_id)
    if md is None or md.test_result_id != result_id:
        raise HTTPException(status_code=404, detail="Mismatch not found")
    updated, status_changed = repo.accept_mismatch(mismatch_id, body.note, body.accepted_by)
    AuditService(db).log(
        request,
        "mismatch.accepted",
        "mismatch",
        mismatch_id,
        {
            "run_id": run_id,
            "result_id": result_id,
            "note": body.note,
            "accepted_by": body.accepted_by,
            "result_status_updated": status_changed,
        },
        actor=body.accepted_by,
    )
    return MismatchAcceptOut(
        id=updated.id,
        accepted=updated.accepted,
        accepted_note=updated.accepted_note,
        accepted_at=updated.accepted_at,
        accepted_by=updated.accepted_by,
        result_status_updated=status_changed,
    )


@router.post("/{run_id}/results/{result_id}/drilldown", response_model=DrilldownOut)
def drilldown_result(
    run_id: str,
    result_id: int,
    payload: DrilldownRequest,
    db: Session = Depends(get_session),
):
    """Re-query both sides grouped by a segment column — live counts.

    Unlike the stored `segment_summary` attached at run completion (see
    ReconciliationEngine._attach_segment_values), this re-runs both sides'
    queries fresh so the caller can catch data drift since the original run.
    """
    from etl_framework.repository.models import TestResult, TestRun, SavedJob
    from api.services.run_executor import RunExecutor
    from api.schemas import RunSettings

    tr = db.get(TestResult, result_id)
    if tr is None or tr.run_id != run_id:
        raise HTTPException(status_code=404, detail="Result not found")
    run = db.query(TestRun).filter(TestRun.run_id == run_id).first()
    if run is None:
        raise HTTPException(status_code=404, detail="Run not found")

    saved = db.query(SavedJob).filter(SavedJob.name == tr.query_name).first()
    if saved is None:
        raise HTTPException(status_code=404, detail=f"Job '{tr.query_name}' not found")
    if saved.job_type != "reconciliation":
        raise HTTPException(status_code=400,
                            detail="Drill-down is only supported for reconciliation jobs")

    snapshot = run.config_snapshot or {}
    ex = RunExecutor(
        db=db, run_id=f"drilldown-{run_id}",
        source_env=run.source_env or "source",
        target_env=run.target_env or "target",
        job_sequence=[],
        run_settings=RunSettings(
            use_live_connections=bool(snapshot.get("source_credentials")),
        ),
        config_snapshot=snapshot,
    )
    job_def = ex._job_to_definition(saved)
    seg = payload.segment_column

    try:
        src_engine, tgt_engine = ex._build_engines(job_def)
        df_src = src_engine.execute_query(job_def.query, job_def.params)
        df_tgt = tgt_engine.execute_query(job_def.query, job_def.params)
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"Drill-down query failed: {exc}")

    def counts(df) -> dict[str, int]:
        if df is None or df.empty or seg not in df.columns:
            return {}
        grouped = df.groupby(df[seg].astype(object).where(df[seg].notna(), "(null)")).size()
        return {str(k): int(v) for k, v in grouped.items()}

    src_counts = counts(df_src)
    tgt_counts = counts(df_tgt)
    if not src_counts and not tgt_counts:
        raise HTTPException(status_code=400,
                            detail=f"Segment column '{seg}' not present in either side")

    rows = []
    for value in sorted(set(src_counts) | set(tgt_counts)):
        s, t = src_counts.get(value, 0), tgt_counts.get(value, 0)
        rows.append(DrilldownRow(value=value, source_count=s, target_count=t, delta=t - s))
    rows.sort(key=lambda r: -abs(r.delta))
    return DrilldownOut(segment_column=seg, job_name=tr.query_name, rows=rows)


@router.get("/trends")
def get_trends(
    job_name: str,
    metric: str = "mismatch_rate",
    window: int = 30,
    db: Session = Depends(get_session),
):
    from datetime import timedelta
    import statistics
    from sqlalchemy import func
    from etl_framework.repository.models import TestRun, TestResult

    signature = (
        db.query(func.count(TestResult.id), func.max(TestResult.id))
        .join(TestRun, TestRun.run_id == TestResult.run_id)
        .filter(TestResult.query_name == job_name)
        .first()
    )
    cache_key = (id(db.bind), job_name, metric, window, signature[0], signature[1])
    cached = _TREND_CACHE.get(cache_key)
    now = time.monotonic()
    if cached and now - cached[0] < _TREND_CACHE_TTL_SECONDS:
        return cached[1]
    if len(_TREND_CACHE) > 500:
        cutoff = now - _TREND_CACHE_TTL_SECONDS
        for k in [k for k, (ts, _) in _TREND_CACHE.items() if ts < cutoff]:
            _TREND_CACHE.pop(k, None)

    cutoff = datetime.now(timezone.utc) - timedelta(days=window)
    rows = (
        db.query(TestResult.value_mismatch_count, TestResult.missing_in_target_count,
                 TestResult.missing_in_source_count, TestResult.source_row_count,
                 TestResult.duration_seconds, TestRun.completed_at)
        .join(TestRun, TestRun.run_id == TestResult.run_id)
        .filter(TestResult.query_name == job_name)
        .filter(TestRun.completed_at.isnot(None))
        .order_by(TestRun.completed_at)
        .all()
    )

    points = []
    for r in rows:
        completed = r.completed_at
        if completed and completed.tzinfo is None:
            completed = completed.replace(tzinfo=timezone.utc)
        if completed and completed < cutoff:
            continue
        total_issues = (r.value_mismatch_count or 0) + (r.missing_in_target_count or 0) + (r.missing_in_source_count or 0)
        src = r.source_row_count or 0
        if metric == "mismatch_rate":
            value = total_issues / src if src else 0.0
        elif metric == "row_count_delta":
            value = float(r.missing_in_target_count or 0) - float(r.missing_in_source_count or 0)
        elif metric == "duration_seconds":
            value = float(r.duration_seconds or 0)
        else:
            value = float(total_issues)
        date_str = completed.strftime("%Y-%m-%d") if completed else "unknown"
        points.append({"date": date_str, "value": round(value, 6)})

    drift_detected = False
    if len(points) >= 3:
        vals = [p["value"] for p in points]
        mean = statistics.mean(vals[:-1])
        try:
            stdev = statistics.stdev(vals[:-1])
            last = vals[-1]
            if stdev > 0:
                drift_detected = (last - mean) > 2 * stdev
            elif last != mean:
                drift_detected = True
        except statistics.StatisticsError:
            pass

    payload = {"job_name": job_name, "metric": metric, "window": window, "points": points, "drift_detected": drift_detected}
    _TREND_CACHE[cache_key] = (now, payload)
    return payload


@router.get("/{run_id}/stream")
async def stream_run(run_id: str, db: Session = Depends(get_session)):
    repo = RunRepository(db)
    if repo.get_run(run_id) is None:
        raise HTTPException(status_code=404, detail="Run not found")

    async def events():
        last_payload = ""
        step_repo = RunStepRepository(db)
        while True:
            db.expire_all()  # force re-fetch; identity map caches stale status without this
            run = repo.get_run(run_id)
            if run is None:
                yield "event: error\ndata: {\"detail\":\"Run not found\"}\n\n"
                break
            total = run.total_tests or 0
            completed = repo.count_completed_results(run_id)
            percent = int(completed / total * 100) if total > 0 else (100 if run.status in _TERMINAL else 0)

            # Determine held/current step from run_steps table
            held_step: int | None = None
            current_step: int | None = None
            steps = step_repo.list_steps(run_id)
            for s in steps:
                if s.status == "HELD":
                    held_step = s.step_index
                    current_step = s.step_index
                    break
                if s.status == "RUNNING":
                    current_step = s.step_index

            payload = {
                "run_id": run.run_id,
                "status": run.status,
                "total_tests": total,
                "completed_tests": completed,
                "current_job": repo.get_current_job(run_id),
                "percent_complete": min(percent, 100),
                "current_step": current_step,
                "held_step": held_step,
            }
            payload_text = json.dumps(payload, default=str)
            if payload_text != last_payload:
                yield f"event: progress\ndata: {payload_text}\n\n"
                last_payload = payload_text
            if run.status in _TERMINAL:
                yield f"event: done\ndata: {payload_text}\n\n"
                break
            await asyncio.sleep(1)

    return StreamingResponse(events(), media_type="text/event-stream")


@router.get("/compare", response_model=RunCompareOut)
def compare_runs(run_a: str, run_b: str, db: Session = Depends(get_session)):
    repo = RunRepository(db)
    ra = repo.get_run(run_a)
    rb = repo.get_run(run_b)
    if ra is None or rb is None:
        raise HTTPException(status_code=404, detail="One or both runs not found")

    def _status_out(r: object) -> RunStatusOut:
        return RunStatusOut(
            run_id=r.run_id, status=r.status,
            started_at=r.started_at, completed_at=r.completed_at,
            total_tests=r.total_tests, passed=r.passed,
            failed=r.failed, slow=r.slow, error=r.error,
            run_type=r.run_type, pair_id=r.pair_id,
        )

    tests_a = {r.query_name: r for r in ra.results}
    tests_b = {r.query_name: r for r in rb.results}
    all_names = sorted(set(tests_a) | set(tests_b))

    improved = regressed = unchanged = only_a = only_b = 0
    tests: list[TestCompareOut] = []
    for name in all_names:
        a = tests_a.get(name)
        b = tests_b.get(name)
        sa = a.effective_status if a else None
        sb = b.effective_status if b else None

        if a and b:
            if sa == "PASSED" and sb != "PASSED":
                regressed += 1
            elif sa != "PASSED" and sb == "PASSED":
                improved += 1
            else:
                unchanged += 1
        elif a:
            only_a += 1
        else:
            only_b += 1

        def _mm(r: object) -> int:
            return (r.value_mismatch_count or 0) + (r.missing_in_target_count or 0) + (r.missing_in_source_count or 0)

        tests.append(TestCompareOut(
            test_name=name,
            status_a=sa,
            status_b=sb,
            duration_a=a.duration_seconds if a else None,
            duration_b=b.duration_seconds if b else None,
            mismatches_a=_mm(a) if a else None,
            mismatches_b=_mm(b) if b else None,
            result_id_a=a.id if a else None,
            result_id_b=b.id if b else None,
        ))

    return RunCompareOut(
        run_a=_status_out(ra),
        run_b=_status_out(rb),
        tests=tests,
        summary={"improved": improved, "regressed": regressed, "unchanged": unchanged,
                 "only_in_a": only_a, "only_in_b": only_b},
    )


@router.get("/{run_id}", response_model=RunDetailOut)
def get_run_detail(run_id: str, db: Session = Depends(get_session)):
    repo = RunRepository(db)
    run = repo.get_run(run_id)
    if run is None:
        raise HTTPException(status_code=404, detail="Run not found")
    results = [_test_result_out(r) for r in run.results]
    return RunDetailOut(
        run_id=run.run_id,
        status=run.status,
        started_at=run.started_at,
        completed_at=run.completed_at,
        total_tests=run.total_tests,
        passed=run.passed,
        failed=run.failed,
        slow=run.slow,
        error=run.error,
        run_type=run.run_type,
        pair_id=run.pair_id,
        source_env=run.source_env,
        target_env=run.target_env,
        config_snapshot=run.config_snapshot,
        results=results,
    )


@router.get("/{run_id}/export")
def export_run_csv(run_id: str, db: Session = Depends(get_session)):
    repo = RunRepository(db)
    run = repo.get_run(run_id)
    if run is None:
        raise HTTPException(status_code=404, detail="Run not found")

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([
        "run_id", "query_name", "status", "effective_status", "agreed_actions",
        "overridden_by", "override_at", "duration_seconds",
        "source_row_count", "target_row_count",
        "value_mismatch_count", "missing_in_target_count", "missing_in_source_count",
        "executed_at",
    ])
    for r in run.results:
        writer.writerow([
            run_id, r.query_name, r.status, r.effective_status, r.override_reason,
            r.override_by, r.override_at,
            r.duration_seconds, r.source_row_count, r.target_row_count,
            r.value_mismatch_count or 0,
            r.missing_in_target_count or 0,
            r.missing_in_source_count or 0,
            r.executed_at,
        ])
    buf.seek(0)
    filename = f"run_{run_id[:8]}_results.csv"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


_STATUS_EMOJI = {"PASSED": "✅", "FAILED": "❌", "ERROR": "❌", "SLOW": "⚠️", "CANCELLED": "⚠️"}


def _sanitize_ci_value(value) -> str:
    text = str(value)
    text = text.replace("\n", " ").replace("\r", " ").replace("`", "'")
    # Neutralize HTML comment delimiters so a malicious ci_context value can't
    # forge a `<!-- ATOM:JOB-STATUS:END -->`-style marker and corrupt a
    # downstream regex-based splice (see Task 6 of the GitLab CI/CD plan).
    return text.replace("<!--", "< !--").replace("-->", "-- >")


def _render_markdown_summary(run) -> str:
    if run.ci_context:
        trigger_line = (
            f"_Last run: {run.completed_at or run.started_at} via GitLab CI "
            f"(commit {_sanitize_ci_value(run.ci_context.get('commit_sha', '?'))}, "
            f"[pipeline]({_sanitize_ci_value(run.ci_context.get('pipeline_url', ''))}), "
            f"ref `{_sanitize_ci_value(run.ci_context.get('ref', '?'))}`)_"
        )
    else:
        trigger_line = f"_Last run: {run.completed_at or run.started_at} (manual)_"

    lines = [
        "## Job Status (auto-updated)",
        "",
        trigger_line,
        "",
        "| Job | Status | Duration |",
        "|-----|--------|----------|",
    ]
    for result in run.results:
        emoji = _STATUS_EMOJI.get(result.effective_status, result.effective_status)
        lines.append(f"| {result.query_name} | {emoji} {result.effective_status} | {result.duration_seconds:.1f}s |")
    lines.append("")
    lines.append(f"[View full run in Atom](/#/runs/{run.run_id})")
    return "\n".join(lines)


@router.get("/{run_id}/markdown-summary", response_class=PlainTextResponse)
def get_run_markdown_summary(run_id: str, db: Session = Depends(get_session)):
    repo = RunRepository(db)
    run = repo.get_run(run_id)
    if run is None:
        raise HTTPException(status_code=404, detail="Run not found")
    return PlainTextResponse(_render_markdown_summary(run), media_type="text/markdown")


@router.delete("/{run_id}", status_code=204)
def delete_run(run_id: str, request: Request, db: Session = Depends(get_session)):
    repo = RunRepository(db)
    if not repo.delete_run(run_id):
        raise HTTPException(status_code=404, detail="Run not found")
    AuditService(db).log(request, "run.deleted", "run", run_id)


# Test result override endpoints
@router.patch("/{run_id}/results/{result_id}/override", response_model=TestResultOut)
def set_test_result_override(
    run_id: str,
    result_id: int,
    body: TestResultOverrideRequest,
    request: Request,
    db: Session = Depends(get_session),
):
    """Set an override on a test result to mark it as passing even when it fails."""
    from etl_framework.repository.models import TestResult

    repo = RunRepository(db)
    if repo.get_run(run_id) is None:
        raise HTTPException(status_code=404, detail="Run not found")

    test_result = db.get(TestResult, result_id)
    if test_result is None or test_result.run_id != run_id:
        raise HTTPException(status_code=404, detail="Test result not found")

    if test_result.status != "FAILED":
        raise HTTPException(
            status_code=409,
            detail="Only failed test results can be passed with agreed actions",
        )

    actor = AuditService.actor_from_request(request) or "unknown"
    test_result.override_status = body.status
    test_result.override_reason = body.reason
    test_result.override_by = actor
    test_result.override_at = datetime.now(timezone.utc)

    db.commit()
    db.refresh(test_result)

    # Log the override action
    AuditService(db).log(
        request,
        "test_result.override_set",
        "test_result",
        result_id,
        {
            "run_id": run_id,
            "query_name": test_result.query_name,
            "original_status": test_result.status,
            "override_status": body.status,
            "reason": body.reason,
            "overridden_by": actor,
        },
        actor=actor,
    )

    return _test_result_out(test_result)


@router.get("/{run_id}/results/{result_id}/override", response_model=dict)
def get_test_result_override(
    run_id: str,
    result_id: int,
    request: Request,
    db: Session = Depends(get_session),
):
    """Get the current override for a test result, if any."""
    from etl_framework.repository.models import TestResult

    repo = RunRepository(db)
    if repo.get_run(run_id) is None:
        raise HTTPException(status_code=404, detail="Run not found")

    test_result = db.get(TestResult, result_id)
    if test_result is None or test_result.run_id != run_id:
        raise HTTPException(status_code=404, detail="Test result not found")

    if test_result.override_status is None:
        return {"override": None}

    return {
        "override": {
            "status": test_result.override_status,
            "reason": test_result.override_reason,
            "overridden_by": test_result.override_by,
            "override_at": test_result.override_at,
        }
    }


@router.delete("/{run_id}/results/{result_id}/override", response_model=TestResultOut)
def delete_test_result_override(
    run_id: str,
    result_id: int,
    request: Request,
    db: Session = Depends(get_session),
):
    """Remove the override from a test result, restoring its original status."""
    from etl_framework.repository.models import TestResult

    repo = RunRepository(db)
    if repo.get_run(run_id) is None:
        raise HTTPException(status_code=404, detail="Run not found")

    test_result = db.get(TestResult, result_id)
    if test_result is None or test_result.run_id != run_id:
        raise HTTPException(status_code=404, detail="Test result not found")

    # Remove the override
    original_status = test_result.status
    test_result.override_status = None
    test_result.override_reason = None
    test_result.override_by = None
    test_result.override_at = None

    db.commit()
    db.refresh(test_result)

    # Log the override removal
    AuditService(db).log(
        request,
        "test_result.override_removed",
        "test_result",
        result_id,
        {
            "run_id": run_id,
            "query_name": test_result.query_name,
            "restored_status": original_status,
        },
    )

    return _test_result_out(test_result)


# ---------------------------------------------------------------------------
# P2 – Badge SVG
# ---------------------------------------------------------------------------

_BADGE_COLORS = {
    "PASSED":    "#4ade80",
    "FAILED":    "#fb7185",
    "SLOW":      "#fbbf24",
    "ERROR":     "#f43f5e",
    "RUNNING":   "#38bdf8",
    "PENDING":   "#94a3b8",
    "COMPLETED": "#4ade80",
}


def _badge_svg(label: str, status: str) -> str:
    color = _BADGE_COLORS.get(status, "#94a3b8")
    label_w = len(label) * 6 + 10
    val_w = len(status) * 7 + 10
    total_w = label_w + val_w
    return (
        f'<svg xmlns="http://www.w3.org/2000/svg" width="{total_w}" height="20">'
        f'<rect width="{label_w}" height="20" fill="#555"/>'
        f'<rect x="{label_w}" width="{val_w}" height="20" fill="{color}"/>'
        f'<text x="{label_w // 2}" y="14" fill="#fff" font-family="DejaVu Sans,Verdana,Geneva,sans-serif" '
        f'font-size="11" text-anchor="middle">{label}</text>'
        f'<text x="{label_w + val_w // 2}" y="14" fill="#fff" font-family="DejaVu Sans,Verdana,Geneva,sans-serif" '
        f'font-size="11" text-anchor="middle">{status}</text>'
        f'</svg>'
    )


from fastapi.responses import Response  # noqa: E402 (local import avoids circular at top)


@router.get("/{run_id}/badge")
def run_badge(run_id: str, db: Session = Depends(get_session)):
    repo = RunRepository(db)
    run = repo.get_run(run_id)
    status = run.status if run else "UNKNOWN"
    svg = _badge_svg("ETL", status)
    return Response(content=svg, media_type="image/svg+xml",
                    headers={"Cache-Control": "no-cache"})


@router.get("/latest/badge")
def latest_badge(job_name: str = "", db: Session = Depends(get_session)):
    from etl_framework.repository.models import TestRun, TestResult
    q = db.query(TestRun).join(TestResult, TestResult.run_id == TestRun.run_id)
    if job_name:
        q = q.filter(TestResult.query_name == job_name)
    run = q.order_by(TestRun.id.desc()).first()
    status = run.status if run else "UNKNOWN"
    svg = _badge_svg(job_name or "ETL", status)
    return Response(content=svg, media_type="image/svg+xml",
                    headers={"Cache-Control": "no-cache"})


# ---------------------------------------------------------------------------
# P2 – Baseline pinning
# ---------------------------------------------------------------------------

@router.post("/{run_id}/set-baseline", status_code=204)
def set_baseline(run_id: str, request: Request, db: Session = Depends(get_session)):
    repo = RunRepository(db)
    run = repo.set_baseline(run_id)
    if run is None:
        raise HTTPException(status_code=404, detail="Run not found")
    AuditService(db).log(request, "run.baseline_set", "run", run_id)


@router.get("/{run_id}/vs-baseline")
def vs_baseline(run_id: str, db: Session = Depends(get_session)):
    from api.schemas import RunCompareOut
    repo = RunRepository(db)
    run = repo.get_run(run_id)
    if run is None:
        raise HTTPException(status_code=404, detail="Run not found")
    baseline = repo.get_baseline(run.source_env or "", run.target_env or "")
    if baseline is None:
        raise HTTPException(status_code=404, detail="No baseline set for this env pair")
    if baseline.run_id == run_id:
        raise HTTPException(status_code=400, detail="Run is already the baseline")
    # Delegate to existing compare logic
    from api.routes.runs import compare_runs
    return compare_runs(run_a=baseline.run_id, run_b=run_id, db=db)


# ---------------------------------------------------------------------------
# P2 – Mismatch distribution
# ---------------------------------------------------------------------------

@router.get("/{run_id}/results/{result_id}/mismatch-distribution")
def mismatch_distribution(run_id: str, result_id: int, top_n: int = 10,
                           db: Session = Depends(get_session)):
    repo = RunRepository(db)
    run = repo.get_run(run_id)
    if run is None:
        raise HTTPException(status_code=404, detail="Run not found")
    return {"result_id": result_id, "distribution": repo.mismatch_distribution(result_id, top_n)}


# ---------------------------------------------------------------------------
# Execution Sequence Scheduler — step endpoints
# ---------------------------------------------------------------------------

@router.get("/{run_id}/steps", response_model=list[RunStepOut])
def list_run_steps(run_id: str, db: Session = Depends(get_session)):
    run = RunRepository(db).get_run(run_id)
    if run is None:
        raise HTTPException(status_code=404, detail="Run not found")
    return RunStepRepository(db).list_steps(run_id)


@router.post(
    "/{run_id}/steps/{step_index}/release",
    response_model=RunStepOut,
)
def release_run_step(
    run_id: str,
    step_index: int,
    body: RunStepReleaseRequest,
    db: Session = Depends(get_session),
):
    run = RunRepository(db).get_run(run_id)
    if run is None:
        raise HTTPException(status_code=404, detail="Run not found")

    step = RunStepRepository(db).get_step(run_id, step_index)
    if step is None:
        raise HTTPException(status_code=404, detail="Step not found")
    if step.status != "HELD":
        raise HTTPException(
            status_code=409,
            detail=f"Step is not held (current status: {step.status})",
        )

    released = RunStepRepository(db).release_step(
        run_id=run_id,
        step_index=step_index,
        action=body.action,
        note=body.note,
        released_by=body.released_by,
    )
    if released is None:
        raise HTTPException(status_code=409, detail="Step could not be released")
    return released
